# -*- coding: utf-8 -*-
import os
from configparser import ConfigParser
from shutil import copyfile

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment

# 存放照片文件夹的目录
# img_root_dir = 'C:\\Users\\ninee\\Desktop'

# 照片文件夹名字列表
# img_dir_name_list = ['中国河南南阳镇平烟草局机房', '中国河南南阳镇平烟草局机房02',
#                      '中国河南南阳镇平烟草局机房03', '中国河南南阳镇平烟草局机房04']

# excel模板文件路径
# excel_file_name = 'C:\\Users\\ninee\\Desktop\\附件2.标准勘察表--基站名.xlsx'

# excel模板文件sheet表名
# sheet_name = '勘察照片'

# excel中照片的标签，也是照片名，供插入照片使用
# label_template = [0, 45, 90, 135, 180, 225, 270, 315,
#                   31, 32, 33, 34, 41, 42, 43, 44,
#                   51, 52, 53, 54, 55, 61, 62, 63, 64,
#                   71, 72, 73, 74, 75, 81, 82, 83,
#                   91, 92, 93, 94, 95, 21, 22, 23, 24, 25, 11]

cfg = ConfigParser()
cfg.read('config.ini', encoding='utf-8')

# img_root_dir = cfg['DEFAULT']['img_root_dir']
# img_dir_name_list = cfg['DEFAULT']['img_dir_name_list'].split()
# excel_file_name = cfg['DEFAULT']['excel_file_name']
# sheet_name = cfg['DEFAULT']['sheet_name']
# label_template = cfg['DEFAULT']['label_template'].split()


def generate_new_excel(img_dir, excel_file_name):
    """
    根据照片文件夹名字img_dir，生成相应excel文件名
    将excel模板文件copy到此新生成的excel文件
    返回新生成excel文件路径
    """
    base_station_name = os.path.basename(img_dir)
    excel_dir_name = os.path.dirname(excel_file_name)
    generate_excel_name = '附件2.标准勘察表--' + base_station_name + '.xlsx'
    generate_excel_path = os.path.join(excel_dir_name, generate_excel_name)

    if os.path.exists(excel_file_name):
        copyfile(excel_file_name, generate_excel_path)

    return generate_excel_path


def load_excel_sheet(excel_path, excel_sheet_name):
    workbook = load_workbook(excel_path, read_only=False)
    worksheet = workbook[excel_sheet_name]
    return worksheet


def get_img_dir_path_list(root_dir, dir_name_list):
    img_dir_path_list = []
    for img_dir_name in dir_name_list:
        img_dir_path = os.path.join(root_dir, img_dir_name)
        img_dir_path_list.append(img_dir_path)
    return img_dir_path_list


def get_images_name(images_dir, images_format='.jpg'):
    """
    遍历照片文件夹，默认获取jpg格式文件名字
    返回{文件名:文件路径}
    """
    images_dict = {}
    for root, dirs, files in os.walk(images_dir):
        # print(root, dirs, files)
        for file in files:
            if os.path.splitext(file)[1] == images_format:
                image_name = os.path.splitext(file)[0]
                image_path = os.path.join(images_dir, file)
                images_dict.update({image_name: image_path})
            else:
                print('在', images_dir, '中:', file, '不是我们需要的', images_format, '格式的文件！')
    return images_dict


def match_coordinate(field_to_match, ws):
    """
    根据文件名，遍历列，再遍历行来匹配excel的坐标，返回匹配到的坐标值
    excel的cell.value 数据类型不固定，由value本身决定
    """
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == field_to_match:
                print(cell.value, '字段在excel中坐标：', cell.coordinate)
                field_coordinate = cell.coordinate
                field_row_num = cell.row
                filed_col_num = cell.column
                return field_coordinate, field_row_num, filed_col_num


def merge_specified_cells(begin_coordinate, end_coordinate, ws):
    """
    合并指定的长*宽区间的cells
    """
    cells_scope = begin_coordinate + ':' + end_coordinate
    ws.merge_cells(cells_scope)


def insert_image(ws, image='img.jpg', image_name='img',
                 cell_coordinate='E3'):
    """
    # 照片高cm/像素高：0.02644
    # 竖照片比例 3:4
    # excel中照片高：7.38cm, 宽：5.57cm
    # excel中像素高：279.12，宽：210.66

    # 横照片比例 4:3
    # excel中照片高：5.24cm, 宽：7.02cm
    # excel中像素高：198.18，宽：265.51
    """
    image_obj = Image(image)
    # print(image_obj.ref, image_obj.anchor, image_obj.format)
    # 11这个照片比较特殊，大小与其他照片不一样
    if image_name == '11':
        if image_obj.height > image_obj.width:
            # 竖照片，resize照片的高宽
            image_size = (382.01, 324.50)
            image_obj.height, image_obj.width = image_size
        else:
            # 横照片，resize照片的高宽
            image_size = (382.01, 525.72)
            image_obj.height, image_obj.width = image_size
    else:
        if image_obj.height > image_obj.width:
            # 竖照片，resize照片的高宽
            image_size = (279.12, 210.66)
            image_obj.height, image_obj.width = image_size
        else:
            # 横照片，resize照片的高宽
            image_size = (198.18, 265.51)
            image_obj.height, image_obj.width = image_size

    # 添加照片
    ws.add_image(image_obj, cell_coordinate)


def cell_alignment(cell_coordinate, ws):
    cell = ws[cell_coordinate]
    cell.alignment = Alignment(horizontal='center', vertical='center')


def cell_alignment_image():
    """
    # VBA macro
    Sub CenterImages(image)
        With ActiveSheet.Shapes(image)
            .Top = Range("B1").Top + (Range("B1").Height - .Height) / 2
            .Left = Range("B1").Left + (Range("B1").Width - .Width) / 2
        End With
    End Sub
    """
    pass


def save_excel(excel_file, wb):
    wb.save(excel_file)


def coordinate_to_scope(row_num, col_num,
                        down_extend_row_num=13, right_extend_col_num=3,
                        up_extend_row_num=0, left_extend_col_num=0):
    """
    根据matched_coordinate找出需要合并的单元格区间，并合并单元格
    matched_coordinate = F17 -> 6,17
    begin_coordinate = E3 -> 5,3
    end_coordinate = H16 -> 8,16
    """
    matched_row_num = row_num
    matched_col_num = col_num

    begin_row_num = matched_row_num + up_extend_row_num
    begin_col_num = matched_col_num + left_extend_col_num
    begin_col_letter = chr(begin_col_num + 64)
    begin_coordinate = begin_col_letter + str(begin_row_num)

    end_row_num = matched_row_num + down_extend_row_num
    end_col_num = matched_col_num + right_extend_col_num
    end_col_letter = chr(end_col_num + 64)
    end_coordinate = end_col_letter + str(end_row_num)

    cells_scope = begin_coordinate + ':' + end_coordinate
    return cells_scope, begin_coordinate, end_coordinate


def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        pass

    try:
        import unicodedata
        unicodedata.numeric(s)
        return True
    except (TypeError, ValueError):
        pass

    return False


def delete_useless_label(label, ws):
    """
    删除没excel中无用的数据标注
    """
    print('--------------------- 删除excel中无用标签 ---------------------------')
    print('excel中标签列表：', label)
    for excel_label in label:
        print('--------------------------------------------------------------')
        no_matched_file_coordinate = match_coordinate(int(excel_label), ws)
        if no_matched_file_coordinate:
            print('删除excel中无用的标签', excel_label)
            ws[no_matched_file_coordinate[0]] = ''
            # 无效标签列表无需删除，使用remove删除
            # 导致列表元素循环时，隔一个循环打印一个
            # label.remove(no_matched_file)
        else:
            print('excel中', excel_label, '标签已经在插入图片后删除，也可能压根没这标签！')


def dir_images_insert_excel(img_dir, label, ws):
    """
    找到文件夹中的文件，获取文件名和文件路径
    在excel中匹配文件名，将匹配到的坐标旁的单元格合并
    将文件路径插入到合并的单元格的开始坐标中
    """
    images_dict = get_images_name(img_dir)
    print('列出所有照片文件：')
    print(images_dict.values())

    for key, value in images_dict.items():
        print('********************** 照片插入操作 ********************************')
        print('要查找的照片名：', key, type(key))
        # 需要图片名为整型
        if is_number(key):
            matched_coordinate_tuple = match_coordinate(int(key), ws)
            print(matched_coordinate_tuple)
            if matched_coordinate_tuple:
                print('插入照片：', value)

                # 匹配到照片后，删除其数据标注
                matched_coordinate = matched_coordinate_tuple[0]
                ws[matched_coordinate] = ''
                # print('目前标签列表：', label)
                # label.remove(int(key))

                matched_row_num = matched_coordinate_tuple[1]
                matched_col_num = matched_coordinate_tuple[2]
                file_path = value
                file_name = key

                # 如果key为11，注意合并单元格的行列数
                if key == '11':
                    coordinate_2_scope = coordinate_to_scope(
                        matched_row_num, matched_col_num, down_extend_row_num=20, right_extend_col_num=11)
                else:
                    coordinate_2_scope = coordinate_to_scope(matched_row_num, matched_col_num)
                cells_scope = coordinate_2_scope[0]
                begin_coordinate = coordinate_2_scope[1]

                # 合并单元格
                ws.merge_cells(cells_scope)
                # 居中（对图片不起作用）
                cell_alignment(begin_coordinate, ws)

                # 将照片文件插入道指定的单元格
                insert_image(ws, image=file_path, image_name=file_name, cell_coordinate=begin_coordinate)
            else:
                print('在excel中，没有找到要插入照片', value, '的位置！')
        else:
            print('图片名不是整数，无法根据图片名去插入到相应的excel label位置！')

    delete_useless_label(label, ws)


def dirs_images_insert_excels(image_dir_name_list, label, sheet_name, excel_file_name):
    """
    遍历所有照片文件夹，分别将其中照片插入到一个相应的excel
    """
    # img_dir_path_list = get_img_dir_path_list(img_root_dir, img_dir_name_list)
    excel_files_list = []
    for images_dir in image_dir_name_list:

        excel_path = generate_new_excel(images_dir, excel_file_name)
        print('@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@')
        print('创建一个excel文件：', excel_path)
        wb = load_workbook(excel_path, read_only=False)
        ws = wb[sheet_name]

        print('################################################################')
        print('开始将', images_dir, '中的照片插入', excel_path)
        # 插入照片到excel
        dir_images_insert_excel(images_dir, label, ws)

        print('==============================================================')
        print('保存excel文件：', excel_path)
        save_excel(excel_path, wb)

        excel_files_list.append(excel_path)

    return excel_files_list

# if __name__ == '__main__':
#     dirs_images_insert_excels(img_dir_name_list, label_template)
